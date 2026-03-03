import sys
from datetime import datetime, time
from pathlib import Path

from openpyxl import load_workbook


def _norm(s):
    return str(s or "").strip()


def _find_sheet(wb, *needles):
    needles = [n.lower() for n in needles if n]
    for name in wb.sheetnames:
        low = name.lower()
        if all(n in low for n in needles):
            return name
    return None


def _header_map(ws, max_rows=8):
    """
    Busca una fila de encabezados en las primeras filas y devuelve {col_name_lower: col_idx}
    """
    for r in range(1, max_rows + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        low = [_norm(v).lower() for v in values]
        if "fecha" in low and "hora" in low and ("celda" in low or "celda id" in low or "celda_id" in low):
            m = {}
            for i, name in enumerate(low, start=1):
                if name:
                    m[name] = i
            return r, m
    # fallback: usar primera fila como header
    r = 1
    values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    low = [_norm(v).lower() for v in values]
    m = {}
    for i, name in enumerate(low, start=1):
        if name:
            m[name] = i
    return r, m


def _header_map_contains(ws, required, max_rows=12):
    """
    Encuentra una fila de encabezados que contenga todas las palabras requeridas (substrings).
    Devuelve (row_idx, {col_name_lower: col_idx})
    """
    req = [str(x).strip().lower() for x in (required or []) if str(x).strip()]
    for r in range(1, max_rows + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        low = [_norm(v).lower() for v in values]
        if all(any(w in col for col in low) for w in req):
            m = {}
            for i, name in enumerate(low, start=1):
                if name:
                    m[name] = i
            return r, m
    return None, {}


def _parse_fecha(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
        return v
    s = _norm(v)
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _parse_hora(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, time):
        return v
    s = _norm(v)
    if not s:
        return None
    # normalizar HH:MM
    if len(s) == 5:
        s = s + ":00"
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except Exception:
            pass
    return None


def _iter_rows(ws, header_row):
    for r in range(header_row + 1, ws.max_row + 1):
        yield r


def main():
    if len(sys.argv) < 2:
        print("Uso: python debug_excel_mapa.py <archivo.xlsx>")
        return 2
    xlsx = Path(sys.argv[1])
    if not xlsx.exists():
        print("No existe:", xlsx)
        return 2

    wb = load_workbook(xlsx, data_only=True)
    print("Archivo:", xlsx.name)
    print("Hojas:", wb.sheetnames)

    res_name = _find_sheet(wb, "resultado", "trafico") or _find_sheet(wb, "resultado") or wb.sheetnames[0]
    dt_name = _find_sheet(wb, "dato", "tecnic") or _find_sheet(wb, "datos", "tecnic") or _find_sheet(wb, "tecnic")

    print("Hoja resultados:", res_name)
    print("Hoja datos técnicos:", dt_name)

    ws = wb[res_name]
    header_row, hmap = _header_map(ws)
    # localizar columnas (flexible)
    def pick(*names):
        for n in names:
            if n in hmap:
                return hmap[n]
        return None

    c_fecha = pick("fecha")
    c_hora = pick("hora")
    c_celda = pick("celda") or pick("celda id") or pick("celda_id")
    print("Header row:", header_row, "cols:", {"fecha": c_fecha, "hora": c_hora, "celda": c_celda})

    best = None  # (dt, row_idx, celda)
    for r in _iter_rows(ws, header_row):
        fv = ws.cell(row=r, column=c_fecha).value if c_fecha else None
        hv = ws.cell(row=r, column=c_hora).value if c_hora else None
        cv = ws.cell(row=r, column=c_celda).value if c_celda else None
        f = _parse_fecha(fv)
        if not f:
            continue
        h = _parse_hora(hv) or time(0, 0, 0)
        dt = datetime.combine(f, h)
        celda = _norm(cv)
        if best is None or dt < best[0]:
            best = (dt, r, celda)

    if not best:
        print("No pude determinar el primer registro (fecha/hora).")
        return 1

    print("Primer registro cronológico:")
    print("- ts:", best[0].isoformat(sep=" "))
    print("- fila:", best[1])
    print("- celda:", best[2] or "—")

    if not dt_name:
        print("No encontré hoja de Datos Técnicos para buscar la celda.")
        return 0

    dt_ws = wb[dt_name]
    # En Datos Técnicos los encabezados suelen ser distintos (Celda, Lat/Long, etc.)
    dt_header_row, dt_hmap = _header_map_contains(dt_ws, required=["celda"])  # mínimo
    if not dt_header_row:
        dt_header_row, dt_hmap = 1, {(_norm(dt_ws.cell(row=1, column=c).value).lower()): c for c in range(1, dt_ws.max_column + 1) if _norm(dt_ws.cell(row=1, column=c).value)}
    # columnas típicas: celda, lat/long
    dt_celda = None
    for k in list(dt_hmap.keys()):
        lk = k.lower()
        if dt_celda is None and ("celda" == lk or lk.startswith("celda")):
            dt_celda = dt_hmap[k]
    # lat/long
    dt_lat = None
    dt_lng = None
    for k in list(dt_hmap.keys()):
        lk = k.lower()
        if dt_lat is None and ("lat" == lk or "latitud" in lk):
            dt_lat = dt_hmap[k]
        if dt_lng is None and (lk in ("long", "lng") or "longitud" in lk):
            dt_lng = dt_hmap[k]
    print("DT header row:", dt_header_row, "cols:", {"celda": dt_celda, "lat": dt_lat, "lng": dt_lng})

    target = (best[2] or "").strip()
    if not target:
        print("El primer registro no tiene celda, no se puede buscar en Datos Técnicos.")
        return 0

    found = []
    for r in range(dt_header_row + 1, dt_ws.max_row + 1):
        cv = _norm(dt_ws.cell(row=r, column=dt_celda).value) if dt_celda else ""
        if cv.strip().upper() == target.upper():
            latv = dt_ws.cell(row=r, column=dt_lat).value if dt_lat else None
            lngv = dt_ws.cell(row=r, column=dt_lng).value if dt_lng else None
            found.append((r, cv, latv, lngv))
            if len(found) >= 5:
                break

    if not found:
        print("NO encontrada en Datos Técnicos:", target)
    else:
        print("Encontrada en Datos Técnicos:", target)
        for r, cv, latv, lngv in found:
            print(f"- fila {r}: celda={cv} lat={latv} lng={lngv}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

